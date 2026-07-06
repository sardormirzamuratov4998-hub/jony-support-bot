from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from bot.database import get_groups_by_support, get_all_supports


DAY_LABELS = {"toq": "Toq kun", "juft": "Juft kun"}


def full_support_card(user, groups) -> str:
    """Admin uchun to'liq karta (telefon raqami bilan)."""
    text = (
        f"👤 <b>{user['first_name']} {user['last_name']}</b>\n"
        f"📞 Telefon: {user['phone'] or '-'}\n"
        f"🏢 Filial: {user['branch'] or '-'}\n\n"
        f"📚 <b>Guruhlari:</b>\n"
    )
    if not groups:
        text += "— Guruhlari yo'q\n"
    else:
        for g in groups:
            day = DAY_LABELS.get(g["day_type"], g["day_type"])
            text += f"• {g['name']} — {g['branch'] or '-'} — {day} — 🕒 {g['time']}\n"
    return text


def examiner_support_card(user, groups) -> str:
    """Examiner uchun karta (telefon raqamisiz)."""
    text = (
        f"👤 <b>{user['first_name']} {user['last_name']}</b>\n"
        f"🏢 Filial: {user['branch'] or '-'}\n\n"
        f"📚 <b>Guruhlari:</b>\n"
    )
    if not groups:
        text += "— Guruhlari yo'q\n"
    else:
        for g in groups:
            day = DAY_LABELS.get(g["day_type"], g["day_type"])
            text += f"• {g['name']} — {day} — 🕒 {g['time']}\n"
    return text


def build_full_excel_report() -> BytesIO:
    """Barcha supportlar bo'yicha filiali va guruhlari tartibida Excel hisobot (raqamsiz)."""
    supports = get_all_supports()

    wb = Workbook()
    ws = wb.active
    ws.title = "Support hisobot"

    headers = ["Filial", "Ism", "Familiya", "Guruh nomi", "Kun turi", "Vaqti"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for user in supports:
        groups = get_groups_by_support(user["id"])
        if not groups:
            ws.append([user["branch"] or "-", user["first_name"], user["last_name"], "-", "-", "-"])
            row_idx += 1
            continue
        for g in groups:
            day = DAY_LABELS.get(g["day_type"], g["day_type"])
            ws.append([
                user["branch"] or "-",
                user["first_name"],
                user["last_name"],
                g["name"],
                day,
                g["time"],
            ])
            row_idx += 1

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
